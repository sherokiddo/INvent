from openpyxl.styles import Border
import openpyxl.styles
from openpyxl import load_workbook
from copy import copy
def Connect():
    try:
        file = load_workbook('methods/shit3.xlsx')
        return file
        print('Подключение успешное ')
    except Exception as ex:
        print('Проблема в подключении ', ex)

file = Connect()

def ShowSheet(file):
    show = file.sheetnames
    return show

def ActiveSheet(file):
    sheet = file.active
    return sheet

def CreateSheet(file,name,sample):
    file.copy_worksheet(file[str(sample)])
    file[str(sample)+' Copy'].title = str(name)
    #file.save('shit2.xlsx')

def ConnectSheet(file,name):
    sheet = file[str(name)]
    return sheet

def AppendValue(sheet,coord,val):
    sheet[coord].value = val
    #print('Данные добавлены')


#CreateSheet(file,'215','Sample')
#sheet = ConnectSheet(file,'269')
#sheet['AE19'].value = 'Стул с Укупником'
#name = ['A','H','O','U','V','W','AD','AK','AR']
#font = openpyxl.styles.Font(name='Times New Roman', size=10)


#file.save('shit3.xlsx')
